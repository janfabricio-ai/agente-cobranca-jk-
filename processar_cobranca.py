"""
Consolidador de Contas a Receber - JK Artes Graficas
Le arquivos da pasta Entrada no Google Drive:
  - zenetti.csv
  - mubys.xls
  - cadastro_clientes.xlsx  (opcional - enriquece telefone)
Gera consolidado de inadimplentes + a vencer
"""

import re
import io
import json
import os
from datetime import datetime, date
from difflib import SequenceMatcher

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# ─────────────────────────────────────────
# CONFIGURACOES
# ─────────────────────────────────────────
BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
CREDENCIAIS_JSON = os.path.join(BASE_DIR, "credenciais_google.json")
DRIVE_IDS_JSON   = os.path.join(BASE_DIR, "drive_ids.json")
SAIDA_DIR        = BASE_DIR
DATA_HOJE        = date.today()

LIMIAR_SIMILARIDADE = 0.82

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets",
]

# ─────────────────────────────────────────
# DRIVE
# ─────────────────────────────────────────
def conectar_drive():
    creds = Credentials.from_service_account_file(CREDENCIAIS_JSON, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)

def baixar_arquivo(service, pasta_id, nome_arquivo):
    """Baixa arquivo da pasta Entrada pelo nome. Retorna bytes ou None."""
    q = f"name='{nome_arquivo}' and '{pasta_id}' in parents and trashed=false"
    res = service.files().list(q=q, fields="files(id,name,modifiedTime)").execute()
    arquivos = res.get("files", [])
    if not arquivos:
        return None, None
    file_id = arquivos[0]["id"]
    modified = arquivos[0].get("modifiedTime", "")
    return _baixar_por_id(service, file_id), modified

def _baixar_por_id(service, file_id):
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf.read()

def procurar_arquivo(service, pasta_id, padrao_regex):
    """
    Busca arquivo na pasta que casa com regex case-insensitive no nome.
    Se houver mais de um, retorna o mais recente (modifiedTime).
    Retorna (bytes, nome_real, modifiedTime) ou (None, None, None).
    """
    q = f"'{pasta_id}' in parents and trashed=false"
    res = service.files().list(
        q=q,
        fields="files(id,name,modifiedTime)",
        pageSize=1000,
    ).execute()
    arquivos = res.get("files", [])
    rx = re.compile(padrao_regex, re.IGNORECASE)
    candidatos = [a for a in arquivos if rx.match(a["name"])]
    if not candidatos:
        return None, None, None
    candidatos.sort(key=lambda a: a.get("modifiedTime", ""), reverse=True)
    escolhido = candidatos[0]
    return _baixar_por_id(service, escolhido["id"]), escolhido["name"], escolhido.get("modifiedTime", "")

# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────
def limpar_valor(texto):
    if not texto:
        return 0.0
    t = re.sub(r'[R$\s]', '', str(texto))
    if not t:
        return 0.0
    if ',' in t:
        # Formato BR (Zenetti CSV): ponto = milhar, virgula = decimal -> "3.060,00"
        t = t.replace('.', '').replace(',', '.')
    # Sem virgula: ja e numero com ponto decimal (Mubisys XLSX, "3060.0") ou
    # inteiro -> NAO remover o ponto, senao infla 10-100x.
    try:
        return float(t)
    except ValueError:
        return 0.0

def similaridade(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def nome_bate(nome_a, nome_b, limiar=LIMIAR_SIMILARIDADE):
    if nome_b.lower().startswith(nome_a.lower()[:20]):
        return True
    if similaridade(nome_a, nome_b) >= limiar:
        return True
    pa = nome_a.lower().split()[:3]
    pb = nome_b.lower().split()[:3]
    if pa == pb:
        return True
    return False

def formatar_moeda(valor):
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def formatar_data(dt):
    if isinstance(dt, date):
        return dt.strftime('%d/%m/%Y')
    return str(dt) if dt else ""

# ─────────────────────────────────────────
# LEITURA ZENETTI
# ─────────────────────────────────────────
def ler_zenetti(conteudo_bytes, apenas_vencidos=True):
    """
    apenas_vencidos=True  -> atraso > 0  (inadimplentes)
    apenas_vencidos=False -> atraso <= 0 (a vencer)
    """
    registros = []
    data_venc_atual = None
    conteudo = conteudo_bytes.decode('latin-1', errors='replace')

    for linha in conteudo.splitlines():
        linha = linha.strip()
        if not linha:
            continue
        m = re.match(r'Vencimento:\s*(\d{2}/\d{2}/\d{4})', linha)
        if m:
            data_venc_atual = datetime.strptime(m.group(1), '%d/%m/%Y').date()
            continue
        if linha.startswith('Documento'):
            continue
        partes = [p.strip() for p in linha.split(';')]
        if len(partes) < 4:
            continue
        documento = partes[0]
        cliente   = partes[2]
        valor     = limpar_valor(partes[3])
        atraso    = int(partes[4].strip()) if len(partes) > 4 and partes[4].strip().lstrip('-').isdigit() else 0

        if not cliente or valor == 0:
            continue

        if apenas_vencidos and atraso <= 0:
            continue
        if not apenas_vencidos and atraso > 0:
            continue

        registros.append({
            'sistema':     'Zenetti',
            'cliente':     cliente,
            'valor':       valor,
            'vencimento':  data_venc_atual,
            'atraso_dias': atraso,
            'documento':   documento,
        })

    return registros

# ─────────────────────────────────────────
# LEITURA MUBYS
# ─────────────────────────────────────────
def _linhas_mubys(conteudo_bytes):
    """
    Extrai as linhas do export Mubisys em formato de lista-de-celulas (strings),
    aceitando os DOIS formatos que o sistema ja exportou:
      - .xls legado = HTML disfarcado (tabela <tr>/<td>)
      - .xlsx real  = Excel OOXML (ZIP, magic 'PK') — formato novo a partir de 06/2026
    O layout de colunas (24 cols, mesmos indices) e identico nos dois.
    """
    # XLSX real comeca com a assinatura ZIP 'PK\x03\x04'
    if conteudo_bytes[:2] == b'PK':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), read_only=True, data_only=True)
        ws = wb.active
        linhas = []
        for row in ws.iter_rows(values_only=True):
            linhas.append(['' if c is None else str(c).strip() for c in row])
        wb.close()
        return linhas

    # Fallback: HTML disfarcado (formato legado .xls)
    conteudo = conteudo_bytes.decode('utf-8-sig', errors='replace')
    linhas_tr = re.findall(r'<tr>(.*?)</tr>', conteudo, re.DOTALL)
    linhas = []
    for tr in linhas_tr:
        tds = re.findall(r'<td[^>]*>(.*?)</td>', tr, re.DOTALL)
        linhas.append([re.sub(r'<[^>]+>', '', c).strip() for c in tds])
    return linhas


def ler_mubys(conteudo_bytes, apenas_vencidos=True):
    """
    apenas_vencidos=True  -> status 'Vencido' (inadimplentes)
    apenas_vencidos=False -> vencimento > hoje (a vencer)
    """
    registros = []
    for tds in _linhas_mubys(conteudo_bytes):
        if len(tds) < 23:
            continue

        def td(i):
            return tds[i] if i < len(tds) else ''

        status   = td(22)
        cliente  = td(0)
        contato  = td(1)
        venc_str = td(12)
        valor    = limpar_valor(td(18))

        if not cliente:
            continue

        try:
            vencimento = datetime.strptime(venc_str, '%d/%m/%Y').date()
        except ValueError:
            vencimento = None

        atraso_dias = (DATA_HOJE - vencimento).days if vencimento else 0

        if apenas_vencidos:
            if 'Vencido' not in status:
                continue
        else:
            # A vencer: vencimento futuro (amanha em diante)
            if vencimento is None or vencimento <= DATA_HOJE:
                continue

        registros.append({
            'sistema':     'Mubys',
            'cliente':     cliente,
            'contato':     contato,
            'valor':       valor,
            'vencimento':  vencimento,
            'atraso_dias': atraso_dias,
            'documento':   td(3),
        })

    return registros

# ─────────────────────────────────────────
# LEITURA CADASTRO
# ─────────────────────────────────────────
def ler_cadastro(conteudo_bytes):
    """
    Le cadastro_clientes.xlsx ou .csv
    Retorna dict: {nome_normalizado -> telefone}
    Colunas esperadas: Nome (ou Cliente), Telefone (ou Fone ou Celular)
    """
    import openpyxl
    cadastro = {}

    # Tenta Excel
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return cadastro

        header = [str(c).strip().lower() if c else '' for c in rows[0]]

        col_nome = col_fone = None
        for i, h in enumerate(header):
            if h in ('nome', 'cliente', 'razao social', 'razão social'):
                col_nome = i
            if h in ('telefone', 'fone', 'celular', 'tel', 'whatsapp'):
                col_fone = i

        if col_nome is None or col_fone is None:
            print(f"  Cadastro: colunas Nome/Telefone nao encontradas. Header: {header}")
            return cadastro

        for row in rows[1:]:
            nome = str(row[col_nome]).strip() if row[col_nome] else ''
            fone = str(row[col_fone]).strip() if row[col_fone] else ''
            if nome and fone and fone.lower() != 'none':
                cadastro[nome] = fone
        return cadastro
    except Exception:
        pass

    # Tenta CSV
    try:
        import csv
        texto = conteudo_bytes.decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(texto), delimiter=';')
        for row in reader:
            keys = {k.lower().strip(): v for k, v in row.items()}
            nome = ''
            fone = ''
            for k in ('nome', 'cliente', 'razao social'):
                if k in keys:
                    nome = keys[k].strip()
                    break
            for k in ('telefone', 'fone', 'celular', 'tel', 'whatsapp'):
                if k in keys:
                    fone = keys[k].strip()
                    break
            if nome and fone:
                cadastro[nome] = fone
        return cadastro
    except Exception as e:
        print(f"  Erro ao ler cadastro: {e}")
        return cadastro

def buscar_fone_cadastro(nome_cliente, cadastro):
    """Tenta casar o nome do cliente com o cadastro."""
    if not cadastro:
        return ''
    for nome_cad, fone in cadastro.items():
        if nome_bate(nome_cliente, nome_cad) or nome_bate(nome_cad, nome_cliente):
            return fone
    return ''

# ─────────────────────────────────────────
# CONSOLIDACAO
# ─────────────────────────────────────────
def consolidar(registros_zenetti, registros_mubys, cadastro=None):
    grupos = {}

    # Indexar Mubys primeiro (nomes completos)
    for reg in registros_mubys:
        nome = reg['cliente']
        if nome not in grupos:
            grupos[nome] = []
        grupos[nome].append(reg)

    # Casar Zenetti com Mubys
    nomes_mubys = list(grupos.keys())
    for reg in registros_zenetti:
        nome_zen = reg['cliente']
        casado = False
        for nome_mubys in nomes_mubys:
            if nome_bate(nome_zen, nome_mubys):
                grupos[nome_mubys].append(reg)
                casado = True
                break
        if not casado:
            if nome_zen not in grupos:
                grupos[nome_zen] = []
            grupos[nome_zen].append(reg)

    resumo = []
    for nome_cliente, regs in grupos.items():
        valor_total  = sum(r['valor'] for r in regs)
        parcelas     = len(regs)
        sistemas     = list(set(r['sistema'] for r in regs))
        mais_antiga  = min((r['vencimento'] for r in regs if r['vencimento']), default=None)
        max_atraso   = max((r['atraso_dias'] for r in regs), default=0)
        contato      = next((r.get('contato', '') for r in regs if r.get('contato')), '')

        # Telefone: prefere contato do Mubys, fallback no cadastro
        telefone = contato
        if not telefone and cadastro:
            telefone = buscar_fone_cadastro(nome_cliente, cadastro)

        resumo.append({
            'cliente':     nome_cliente,
            'valor_total': valor_total,
            'parcelas':    parcelas,
            'sistemas':    ' + '.join(sorted(sistemas)),
            'mais_antiga': mais_antiga,
            'max_atraso':  max_atraso,
            'telefone':    telefone,
            'detalhes':    regs,
        })

    # Ordenacao: prioridade de urgencia, depois valor decrescente
    def urgencia_key(r):
        d = r['max_atraso']
        if d > 90:   return 0
        if d > 60:   return 1
        if d > 30:   return 2
        return 3

    resumo.sort(key=lambda x: (urgencia_key(x), -x['valor_total']))
    return resumo


def consolidar_a_vencer(reg_zen, reg_mub, cadastro=None):
    """Consolida registros a vencer (mesma logica, mas sem atraso)."""
    grupos = {}

    for reg in reg_mub:
        nome = reg['cliente']
        if nome not in grupos:
            grupos[nome] = []
        grupos[nome].append(reg)

    nomes_mubys = list(grupos.keys())
    for reg in reg_zen:
        nome_zen = reg['cliente']
        casado = False
        for nome_mubys in nomes_mubys:
            if nome_bate(nome_zen, nome_mubys):
                grupos[nome_mubys].append(reg)
                casado = True
                break
        if not casado:
            if nome_zen not in grupos:
                grupos[nome_zen] = []
            grupos[nome_zen].append(reg)

    resumo = []
    for nome_cliente, regs in grupos.items():
        valor_total = sum(r['valor'] for r in regs)
        parcelas    = len(regs)
        sistemas    = list(set(r['sistema'] for r in regs))
        mais_prox   = min((r['vencimento'] for r in regs if r['vencimento']), default=None)
        contato     = next((r.get('contato', '') for r in regs if r.get('contato')), '')
        telefone    = contato or (buscar_fone_cadastro(nome_cliente, cadastro) if cadastro else '')

        resumo.append({
            'cliente':      nome_cliente,
            'valor_total':  valor_total,
            'parcelas':     parcelas,
            'sistemas':     ' + '.join(sorted(sistemas)),
            'mais_proxima': mais_prox,
            'telefone':     telefone,
            'detalhes':     regs,
        })

    resumo.sort(key=lambda x: (x['mais_proxima'] or date.max, -x['valor_total']))
    return resumo


# ─────────────────────────────────────────
# MAIN — retorna dados para uso externo
# ─────────────────────────────────────────
def carregar_dados():
    """
    Baixa arquivos do Drive, processa e retorna todos os dados.
    Retorna: (resumo_vencidos, resumo_a_vencer, zen_venc, mub_venc)
    """
    with open(DRIVE_IDS_JSON) as f:
        ids = json.load(f)
    pasta_entrada = ids["pasta_entrada"]

    print("Conectando ao Drive...")
    service = conectar_drive()

    print("Procurando zenetti (csv)...")
    # Aceita nome legado 'zenetti*.csv' OU nome nativo do export Zenetti
    # ('Contas a receber por vencimento.csv')
    zen_bytes, zen_nome, _ = procurar_arquivo(
        service, pasta_entrada,
        r'^(zenetti|contas[\s_-]*a[\s_-]*receber).*\.csv$'
    )
    if not zen_bytes:
        raise FileNotFoundError("zenetti.csv nao encontrado na pasta Entrada do Drive")
    print(f"  encontrado: {zen_nome}")

    print("Procurando mubys/mubisys (xls/xlsx)...")
    # Aceita nome legado 'mub*.xls' OU nome nativo do export Mubisys
    # ('Resultado_DD_MM_YYYY.xls')
    mub_bytes, mub_nome, _ = procurar_arquivo(
        service, pasta_entrada,
        r'^(mub[iy]?s|resultado[_\s-]*\d).*\.xlsx?$'
    )
    if not mub_bytes:
        raise FileNotFoundError("mubys.xls nao encontrado na pasta Entrada do Drive")
    print(f"  encontrado: {mub_nome}")

    print("Procurando cadastro_clientes (opcional)...")
    cad_bytes, cad_nome, _ = procurar_arquivo(
        service, pasta_entrada, r'^cadastro[_ ]?clientes.*\.(xlsx|csv)$'
    )
    cadastro = {}
    if cad_bytes:
        cadastro = ler_cadastro(cad_bytes)
        print(f"  encontrado: {cad_nome} ({len(cadastro)} clientes)")
    else:
        print("  Cadastro nao encontrado — telefones apenas do Mubys")

    print("\nProcessando vencidos...")
    zen_venc = ler_zenetti(zen_bytes, apenas_vencidos=True)
    mub_venc = ler_mubys(mub_bytes, apenas_vencidos=True)
    print(f"  Zenetti: {len(zen_venc)} parcelas | Mubys: {len(mub_venc)} parcelas")
    resumo_vencidos = consolidar(zen_venc, mub_venc, cadastro)
    print(f"  {len(resumo_vencidos)} clientes inadimplentes")

    print("\nProcessando a vencer...")
    zen_av = ler_zenetti(zen_bytes, apenas_vencidos=False)
    mub_av = ler_mubys(mub_bytes, apenas_vencidos=False)
    print(f"  Zenetti: {len(zen_av)} parcelas | Mubys: {len(mub_av)} parcelas")
    resumo_av = consolidar_a_vencer(zen_av, mub_av, cadastro)
    print(f"  {len(resumo_av)} clientes com parcelas a vencer")

    return resumo_vencidos, resumo_av, zen_venc, mub_venc


if __name__ == '__main__':
    resumo_v, resumo_av, zen, mub = carregar_dados()
    total_v  = sum(r['valor_total'] for r in resumo_v)
    total_av = sum(r['valor_total'] for r in resumo_av)
    print(f"\nTotal vencido:   {formatar_moeda(total_v)}")
    print(f"Total a vencer:  {formatar_moeda(total_av)}")
    print(f"Total geral:     {formatar_moeda(total_v + total_av)}")
